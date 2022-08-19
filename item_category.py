import requests 
from bs4 import BeautifulSoup
import openpyxl as op
import urllib.request
import pandas as pd

# cid(카테고리 id) 정의
toCid = {'수납/정리': '019000000000000',
        '주방/욕실/청소': '020000000000000',
        '가구/인테리어': '022000000000000',
        '사무/문구/디지털': '021000000000000',
        '가전/레져/식품': '023000000000000',
        '키즈/뷰티/패션잡화': '024000000000000',
        '다이소 매장상품': '010000000000000',
        '포장재 전문관': '027000000000000'}

def write_data(input, output, imgPath, sheet):

    input_wb = op.load_workbook(input)  # 입력을 읽어올 엑셀파일 열기
    input_ws = input_wb[sheet]          # 시트 지정
    row_max = input_ws.max_row          # 최대행값 저장

    output_wb = op.load_workbook(output)  # 결과를 저장할 엑셀파일 열기
    output_ws = output_wb[sheet]          # 시트 지정

    for r in range(2, row_max+1):  # input 파일의 2행부터 마지막행까지 탐색
        categories = []

        type = str(input_ws.cell(row=r, column=1).value)  # 물품분류
        if type == 'None':
            type = str(input_ws.cell(row=r-1, column=1).value)

        search = str(input_ws.cell(row=r, column=2).value)  # 검색어
        if search == 'None':
            continue

        search_include = str(input_ws.cell(row=r, column=3).value).split(', ')    # 포함단어
        search_except = str(input_ws.cell(row=r, column=4).value).split(', ')     # 제외단어

        for c in range(7, input_ws.max_column+1):  # 카테고리명을 리스트에 저장
            category = str(input_ws.cell(r, c).value)
            if category == 'None':
                continue
            categories.append(category)

        print("-------------" + search + "-------------")

        num = 1  # 순번
        for category in categories:  # 리스트에 저장해놨던 카테고리를 하나씩 탐색

            index1 = category.find('(')
            index2 = category.find(')')
            itemNum = int(category[index1+1:index2].replace(',', '')) # 현재 카테고리에 속한 물품 개수

            categoryPage = int(itemNum/50) + 1 # 물품 개수를 이용해 총 페이지 수 구하기 (한 페이지당 물품 50개)
            if categoryPage > 150:  # 150 페이지가 넘어가는 경우 150 페이지까지만 탐색
                categoryPage = 150
            category = category[:index1]

            try:
                cid = toCid[category]  # 한글로 된 카테고리명을 url에서 쓰이는 cid로 변환
            except:
                print(category + "---unknown category!")
                continue

            print("-------------" + category + "-------------")

            for page in range(1, categoryPage+1):  # 1페이지부터 마지막 페이지까지 반복
                print(page)

                # 페이지 url에서 html 소스 읽어옴
                url = "https://www.daisomall.co.kr/shop/search.php?nset=1&page={}&max=50&search_text={}&orderby=daiso_ranking1&cid={}&depth=1".format(page, search, cid)
                res = requests.get(url, headers={'User-Agent':'Mozilla/5.0'})
                res.raise_for_status()
                soup = BeautifulSoup(res.text, "lxml")  # html 소스를 soup 객체에 담음

                # 해당 페이지에 나와 있는 상품을 "items" 리스트에 모두 저장
                items = soup.find_all("li", {"class": "float01 search_goods_list"})

                for item in items:  # 상품을 하나씩 탐색

                    title = item.find('div', {"style": "margin-top:10px;height:38px;"})
                    itemName = title.find('a').get("title")  # 상품명

                    flag = 0
                    if itemName.find("<b>") == -1:  # 상품명에 검색어가 포함되지 않은 항목 제외
                        for word in search_include: # 검색어가 포함되지 않았으나 '포함단어' 리스트에 있는 단어를 포함한 경우 제외하지 않음
                            if itemName.find(word) != -1:
                                flag = 1
                        if flag == 0:
                            continue
                    
                    itemName = itemName.replace("<b>", '')
                    itemName = itemName.replace("</b>", '')

                    # 모든 상품에 일괄적으로 제외할 단어들
                    if itemName.find("밀크북") != -1:
                        continue
                    if itemName.find("양장") != -1:
                        continue
                    
                    # '제외단어' 리스트에 있는 단어가 하나라도 포함된 경우 제외
                    flag = 0
                    for word in search_except:
                        if itemName.find(word) != -1:
                            flag = 1
                    if flag == 1:
                        continue

                    print(itemName)

                    price = item.find('div', {"style": "margin-top:12px;"})  # 상품가격
                    itemPrice = price.find('strong').text
                    itemPrice = itemPrice.replace("원", '') # "원" 과 쉼표 지우고 숫자만 남기기
                    itemPrice = itemPrice.replace(",", '')

                    img = item.find('div', {"class": "goods_line_img"})  
                    imgUrl = img.find('img').get('src')       # 상품 이미지 url
                    imgName = "{}{}.jpg".format(search, num)  # 상품 이미지 파일 이름 형식

                    # 상품 이미지 다운로드
                    urllib.request.urlretrieve(imgUrl, imgPath+imgName)

                    itemId = item.find('a').get('href')  # 상품번호
                    itemId = itemId[24:34]

                    # 상품번호로 해당 상품의 상세페이지 url 접속해 html 소스 읽어옴
                    itemUrl = "https://www.daisomall.co.kr/shop/goods_view.php?id={}&depth=1&search_text={}".format(itemId, search)

                    # 가끔씩 다이소몰 서버 측에서 request error가 나는 경우가 있어 예외 처리 해줌
                    try:
                        itemRes = requests.get(itemUrl, headers={'User-Agent':'Mozilla/5.0'})
                    except:
                        print("request error!")
                        continue
                    itemRes.raise_for_status()
                    soupItem = BeautifulSoup(itemRes.text, "lxml")

                    # 해당 상품이 '투데이 딜'이라는 이벤트에 포함되면 상품 페이지 html이 달라짐
                    # 이 경우 상품번호를 읽어오는 방식이 달라지고 가격도 변동되기 때문에 예외 처리해 해당 상품을 skip
                    try:
                        itemNum = soupItem.find('td', {"class": "color_63 line_h160"}).find('strong').text 
                    except:
                        print("itemNum error!")
                        continue
                    
                    itemCategories = []
                    for itemCategory in soupItem.find_all('option', {"selected": ""}):  # 해당 상품의 세부 카테고리를 itemCategories 리스트에 저장
                        if str(itemCategory).find("selected") == -1:
                            continue
                        itemCategory = itemCategory.get_text()
                        itemCategories.append(itemCategory)

                    try: # 엑셀에 한 행을 통째로 쓰기
                        data = [type, None, search, num, int(itemNum), itemCategories[0]+">"+itemCategories[1]+">"+itemCategories[2], itemName, imgUrl, int(itemPrice), itemUrl]
                    except:
                        print("category error!")  # 가끔씩 세부 카테고리 3개 중 2개만 있는 경우를 위한 예외 처리
                        continue
                    output_ws.append(data)
                    num+=1
                    
            output_wb.save(output)


# 물품코드 로딩
def load_code(input, output, sheet):

    wb = op.load_workbook(output) # 결과를 저장할 엑셀 파일 열기
    ws = wb.active                # 시트 지정
    code_wb = op.load_workbook(input)  # 물품 코드를 읽어올 입력 파일 열기
    code_ws = code_wb['물품코드표']     # 입력 파일 중 물품 코드가 있는 시트 지정

    row_max = ws.max_row  #                 
    code_row_max = code_ws.max_row
    
    codeDic = {'item': 'code'}
    for r in range(2, code_row_max+1):
        code = str(code_ws.cell(row=r, column=1).value)
        item = str(code_ws.cell(row=r, column=2).value)
        desc = str(code_ws.cell(row=r, column=3).value)
        if desc.find("삭제") != -1:
            continue
        codeDic[item] = code

    for r in range(2, row_max+1):
        temp = str(ws.cell(row=r, column=1).value)
        try:
            ws.cell(row=r, column=2).value = int(codeDic[temp])
        except:
            print("Key error! --- " + temp)
            continue
        
    wb.save(output)

# 중복 행 제거
def drop_duplicates(filename_xlsx):
    df = pd.read_excel(filename_xlsx, engine='openpyxl')  # 엑셀파일 읽어오기
    df = df.drop_duplicates(subset='상품번호')  # 중복 행 제거
    df.to_excel(filename_xlsx[:-5]+"_dptest.xlsx", index=False)

# main 함수
if __name__ == "__main__":

    sheet = 'sheet_name'
    input = "input_file_name.xlsx"    # 읽어올 물품 리스트
    output = "output_file_name.xlsx"  # 결과를 저장할 xlsx 파일 이름
    imgPath = "imgFolder_name/"            # 이미지 파일이 저장될 경로
    columns_name = ["물품분류", "물품코드", "물품종", "순번", "상품번호", "카테고리", "상품명", "상품사진", "가격", "링크"]  # 컬럼명 지정

    # # #--- 새 엑셀파일 생성 시
    # output_wb = op.Workbook()
    # output_ws = output_wb.create_sheet(sheet)
    # output_ws.append(columns_name)
    # output_wb.save(output)

    # #--- 기존 엑셀파일에 추가 시
    # output_wb = op.load_workbook(output)  # 결과를 저장할 엑셀파일
    # output_ws = output_wb.create_sheet(sheet)
    # output_ws.append(columns_name)
    # output_wb.save(output)

    # write_data(input, output, imgPath, sheet)
    # drop_duplicates(output)
    load_code(input, output, sheet) # !!!!!! 주석처리 확인 !!!!!!