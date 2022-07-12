import openpyxl as op
import requests 
from bs4 import BeautifulSoup

# 각 물품별 카테고리 현황 업로드 함수
def loadCategory(excel_path, result_file):
    wb = op.load_workbook(excel_path) # 엑셀파일 열기
    ws = wb.active
 
    row_max = ws.max_row # 최대행값 저장

    for r in range(2, row_max+1):
        search = str(ws.cell(row=r, column=2).value)
        if search == "None":
            continue
        print(search)
        url = "https://www.daisomall.co.kr/shop/search.php?nset=1&max=50&search_text={}&orderby=daiso_ranking1".format(search)
        
        res = requests.get(url)
        res.raise_for_status()
        soup = BeautifulSoup(res.text, "lxml")

        try:
            max_item = int(soup.find("span", {"class": "font_normal size_16"}).text)
        except:
            max_item = 0

        categories = []
        for categoryBox in soup.find_all("li", {"class":"float01"}):
            try:
                category = str(categoryBox.find('a').text)[1:]
                categories.append(category)
            except:
                break
        
        ws.cell(row=r, column=6).value = "총 상품 " + str(max_item) + "개"
        for i in range(0, len(categories)):
            ws.cell(row=r, column=i+7).value = categories[i]

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20    
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 10
    for i in range(6, 17):
        ws.column_dimensions[chr(67+i)].width = 15
        
    wb.save(result_file)

if __name__ == "__main__":

    excel_path = "item_list.xlsx"               # 원본 엑셀파일
    result_file = "item_category_list.xlsx"     # 결과를 저장할 엑셀 파일 이름

    loadCategory(excel_path, result_file)