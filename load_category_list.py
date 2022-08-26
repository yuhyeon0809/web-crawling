import openpyxl as op
import requests 
from bs4 import BeautifulSoup

# 각 물품별 카테고리 현황 업로드 함수
def loadCategory(input, output, sheet):
    wb = op.load_workbook(input) # 엑셀파일 열기
    ws = wb[sheet] # 시트 지정
 
    row_max = ws.max_row # 최대행값 저장

    for r in range(2, row_max+1): # 2행부터 마지막행까지 반복
        search = str(ws.cell(row=r, column=2).value)  # '검색어' 열의 데이터를 search 변수에 저장
        if search == "None":  # '검색어' 열이 빈 칸일 경우 검색하지 않음
            continue
        print(search)
        url = "https://www.daisomall.co.kr/shop/search.php?nset=1&max=50&search_text={}&orderby=daiso_ranking1".format(search)  # 해당 검색어의 검색 결과 페이지 url
        
        res = requests.get(url)  # url에 http 요청을 보낸 후 응답 전문을 res 객체에 담음
        res.raise_for_status()   # 요청/응답 코드의 status가 200이 아니면 예외를 발생시킴 (멈추고 에러가 났다는 것을 알려줌)
        soup = BeautifulSoup(res.text, "lxml")  # http 응답으로 받은 html 소스를 BeautifulSoup 객체(soup)에 저장

        try:
            max_item = int(soup.find("span", {"class": "font_normal size_16"}).text)  # 총 상품 개수
        except:
            max_item = 0

        categories = []  # 해당 검색어의 상품들이 속한 모든 카테고리를 담을 빈 리스트 생성
        for categoryBox in soup.find_all("li", {"class":"float01"}):  # 얻은 카테고리 정보를 categories 리스트에 저장
            try:
                category = str(categoryBox.find('a').text)[1:]
                categories.append(category)
            except:
                break
        
        ws.cell(row=r, column=6).value = "총 상품 " + str(max_item) + "개"  # 엑셀 파일의 현재 행(r)의 '총상품개수'열(6번째 열)에 총 상품 개수를 입력
        for i in range(0, len(categories)):
            ws.cell(row=r, column=i+7).value = categories[i]  # 엑셀 파일의 현재 행(r) '카테고리'열 (7번째 열) 부터 카테고리를 하나씩 입력

    # 열 너비 조정
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20    
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 10
    for i in range(6, 17):
        ws.column_dimensions[chr(67+i)].width = 15
        
    wb.save(output)  # 결과를 output에 지정해뒀던 이름으로 저장

if __name__ == "__main__":

    sheet = "sheet_name"                         # 대상 시트명
    input = "input_file_name.xlsx"               # 원본 엑셀파일 이름
    output = "output_file_name.xlsx"             # 결과를 저장할 엑셀 파일 이름

    loadCategory(input, output, sheet)