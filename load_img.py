import openpyxl as op
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

# 다운받은 물품 이미지를 엑셀파일에 업로드
def loadImg(excel_path, image_path, result_file):
    wb = op.load_workbook(excel_path) # 엑셀파일 열기
    ws = wb.active
 
    row_max = ws.max_row # 최대행값 저장

    for r in range(2, row_max+1):               # 2행부터 마지막행까지 반복
        search = str(ws.cell(row=r, column=3).value)
        num = str(ws.cell(row=r, column=4).value)
        file_name = search + num + ".jpg"       # 이미지 파일 이름 (물품종(검색어)+순번.jpg)
        try:
            img = Image(image_path + file_name) # 이미지 파일 객체화
        except:
            continue
        img.width = 128                         # 이미지 크기 조정
        img.height = 130
        ws.add_image(img, "H"+str(r))           # H(r) 셀에 이미지 삽입
      
        ws.row_dimensions[r].height = 100       # 행 높이 조정
        ws.alignment = Alignment(horizontal='center', vertical='center')

    # 열 너비 조정
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['E'].width = 10.4
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 50    
    ws.column_dimensions['H'].width = 15.4
    ws.column_dimensions['J'].width = 15    

    wb.save(result_file)   # 결과 파일 저장

# main 함수
if __name__ == "__main__":

    sheet_name = "sheet_name"
    input = "test.xlsx"               # 텍스트 엑셀파일 이름
    output = "test_result_img.xlsx"   # 이미지 로딩한 결과를 저장할 파일 이름
    image_path = "imgFolder_path/"    # 이미지 파일 경로

    loadImg(input, image_path, output)