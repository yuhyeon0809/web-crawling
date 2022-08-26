import requests 
from bs4 import BeautifulSoup
import openpyxl as op

# 상품 옵셥별로 분리해 저장하는 함수
def load_option(input, output):
    input_wb = op.load_workbook(input)    # 입력 파일 열기
    input_ws = input_wb.active
    row_max = input_ws.max_row # 최대행값

    output_wb = op.load_workbook(output)  # 결과 파일 열기
    output_ws = output_wb.active
    
    for r in range(2, row_max+1): # 2행부터 마지막행까지 반복
        row_data = []  # input 파일의 현재 행 데이터를 담을 빈 리스트 생성
        for i in range(1, 11): # input 파일의 현재 행 데이터를 row_data 리스트에 저장
            row_data.append(input_ws.cell(row=r, column=i).value)

        url = row_data[9]  # 상품 url
        try:  # 옵션 정보를 스크래핑하는 과정에서 에러가 날 경우를 대비한 예외처리
            res = requests.get(url, headers={'User-Agent':'Mozilla/5.0'})  # url로 접속해 http 소스 받아옴
            res.raise_for_status()
            soup = BeautifulSoup(res.text, "lxml")       # html 소스를 BeautifulSoup 객체에 저장
            optionBox = soup.find("select", {"class": "goods_options required_option"})  # 상품의 옵션이 있는 부분의 html 소스를 optionBox에 저장
            options = []                                 # 해당 상품의 옵션을 담을 빈 리스트 생성
            for option in optionBox.find_all('option'):  # 상품의 옵션들을 options 리스트에 저장                   
                text = option.text
                options.append(text)
        except:
            output_ws.append(row_data)  
            output_wb.save(output)
            continue

        if len(options) == 0:  # 옵션이 없는 상품의 경우 옵션 열을 비워두고 그대로 저장
            output_ws.append(row_data)
            output_wb.save(output)
            continue
        
        origin_price = row_data[8]  # 옵션 선택으로 인한 추가 비용 합산 전 원래 가격
        try:
            int(origin_price)
        except:
            output_ws.append(row_data)
            output_wb.save(output)
            continue

        for i in range(1, len(options)): # 저장해둔 옵션들을 하나씩 탐색
            add_price = 0
            if options[i].find('품절') != -1:  # 품절상품 제외
                continue

            if options[i].find('+') != -1:  # 추가 비용이 붙는 옵션의 경우
                index_1 = options[i].find('+')
                index_2 = options[i].find('원')
                add_price = options[i][index_1:index_2].replace(',', '')
                try:
                    row_data[8] = origin_price + int(add_price)  # 원래 가격에 추가 비용을 합산
                except:
                    continue
            if options[i].find('-') != -1:  # 가격이 절감되는 옵션의 경우
                index_1 = options[i].find('-')
                index_2 = options[i].find('원')
                add_price = options[i][index_1:index_2].replace(',', '')
                try:
                    row_data[8] = origin_price - int(add_price)  # 원래 가격에 절감되는 비용을 빼줌
                except:
                    continue

            row_data.append(options[i])  # 현재 행의 데이터가 담긴 리스트에 옵션 정보 추가
            output_ws.append(row_data)   # 출력파일에 현재 행 데이터 쓰기 
            print(str(row_data[6]) + '------------' + options[i])

            row_data[8] = origin_price   # 변동되었던 가격 복구
            row_data.remove(options[i])  # 다음 옵션을 위해 옵션 칸 비우기

        output_wb.save(output) # 저~장~


if __name__ == "__main__":

    input = "input.xlsx"           # 입력 파일 이름
    output = "output.xlsx"         # 결과 파일 이름      
    columns_name = ["물품분류", "물품코드", "물품종", "순번", "상품번호", "카테고리", "상품명", "상품사진", "가격", "링크", "비고"] # 컬럼명 지정

    # --- 새 엑셀파일 생성 시 --- #
    output_wb = op.Workbook()
    output_ws = output_wb.active
    output_ws.append(columns_name)
    output_wb.save(output)
    
    load_option(input, output)