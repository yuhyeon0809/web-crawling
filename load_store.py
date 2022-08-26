import requests 
from bs4 import BeautifulSoup
import openpyxl as op

def load_store(input, output):
    input_wb = op.load_workbook(input)
    input_ws = input_wb.active
    row_max = input_ws.max_row # 최대행값

    output_wb = op.load_workbook(output)
    output_ws = output_wb.active
    
    for r in range(40001, row_max+1): # 5행부터 마지막행까지 반복

        row_data = [None]
        for i in range(2, 11): # 입력 파일의 현재 행 데이터를 row_data 리스트에 저장
            row_data.append(input_ws.cell(row=r, column=i).value)

        url = input_ws.cell(row=r, column=10).value

        try:
            res = requests.get(url, headers={'User-Agent':'Mozilla/5.0'})  # url로 접속해 html 소스 받아옴
            res.raise_for_status()
            soup = BeautifulSoup(res.text, "lxml")
            sellerBox = soup.find("dl", {"class": "minishop_seller_info"})
            seller = sellerBox.find("span", {"class": "vm"}).text
        except:
            continue
        
        row_data.append(None)
        row_data.append(seller)

        print(str(row_data[1]) + str(input_ws.cell(row=r, column=7).value) + "-----------------------" + seller)

        output_ws.append(row_data)
        output_wb.save(output) 


if __name__ == "__main__":

    input = "스토어.xlsx"           # 입력 파일 이름
    output = "스토어result5.xlsx"     # 출력 파일 이름      
    columns_name = [None, "순번", "물품분류", "물품코드", "상품번호", "카테고리", "상품명", "가격", "구매여부", "링크", "옵션 선택", "스토어"] # 컬럼명 지정

    # --- 새 엑셀파일 생성 시 --- #
    output_wb = op.Workbook()
    output_ws = output_wb.active
    output_ws.append(columns_name)
    output_wb.save(output)
    
    load_store(input, output)