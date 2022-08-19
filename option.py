import requests 
from bs4 import BeautifulSoup
import openpyxl as op

def load_option(input, output):
    input_wb = op.load_workbook(input)
    input_ws = input_wb.active
    row_max = input_ws.max_row # 최대행값

    output_wb = op.load_workbook(output)
    output_ws = output_wb.active
    
    for r in range(2069, row_max+1): # 2행부터 마지막행까지 반복
        row_data = []
        for i in range(1, 11): # 입력 파일의 현재 행 데이터를 row_data 리스트에 저장
            row_data.append(input_ws.cell(row=r, column=i).value)

        url = row_data[9]  # 상품 url
        try:
            res = requests.get(url, headers={'User-Agent':'Mozilla/5.0'})  # url로 접속해 http 소스 받아옴
            res.raise_for_status()
            soup = BeautifulSoup(res.text, "lxml")
            optionBox = soup.find("select", {"class": "goods_options required_option"})
            options = []
            for option in optionBox.find_all('option'):  # 상품의 옵션들을 options 리스트에 저장                   
                text = option.text
                options.append(text)
        except:
            output_ws.append(row_data)
            output_wb.save(output)
            continue

        if len(options) == 0:  
            output_ws.append(row_data)
            output_wb.save(output)
            continue
        
        origin_price = row_data[8]  # 옵션 선택으로 인한 추가 비용 합산 전 원래 가격
        try:
            int(origin_price)
        except:
            output_ws.append(row_data)
            output_wb.save(output)
            continue

        for i in range(1, len(options)): # 저장해둔 옵션들을 하나씩 탐색
            add_price = 0
            if options[i].find('품절') != -1:  # 품절상품 제외
                continue

            if options[i].find('+') != -1:  # 추가 비용이 붙는 옵션의 경우
                index_1 = options[i].find('+')
                index_2 = options[i].find('원')
                add_price = options[i][index_1:index_2].replace(',', '')
                try:
                    row_data[8] = origin_price + int(add_price)
                except:
                    continue
            if options[i].find('-') != -1:  # 가격이 절감되는 옵션의 경우
                index_1 = options[i].find('-')
                index_2 = options[i].find('원')
                add_price = options[i][index_1:index_2].replace(',', '')
                try:
                    row_data[8] = origin_price - int(add_price)
                except:
                    continue

            row_data.append(options[i]) 
            output_ws.append(row_data)  # 출력파일에 현재 행 데이터 쓰기 
            print(str(row_data[6]) + '------------' + options[i])

            row_data[8] = origin_price   # 변동되었던 가격 복구
            row_data.remove(options[i])  # 다음 옵션을 위해 옵션 칸 비우기

        output_wb.save(output) # 저~장~


if __name__ == "__main__":

    input = "주방용품.xlsx"           # 입력 파일 이름
    output = "주방용품_옵션2.xlsx"     # 출력 파일 이름      
    columns_name = ["물품분류", "물품코드", "물품종", "순번", "상품번호", "카테고리", "상품명", "상품사진", "가격", "링크", "비고"] # 컬럼명 지정

    # --- 새 엑셀파일 생성 시 --- #
    output_wb = op.Workbook()
    output_ws = output_wb.active
    output_ws.append(columns_name)
    output_wb.save(output)
    
    load_option(input, output)