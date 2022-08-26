from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from time import sleep
import openpyxl as op

input = "촬영 대상 물품 분류체계_v1.8_권혁진.xlsx"

driver = webdriver.Chrome()

wb = op.load_workbook(input)
ws = wb.active
row_max = ws.max_row

error = []

def test_one():
    url = "https://www.daisomall.co.kr/shop/goods_view.php?id=0005343781&cid=&depth=&search_text=%EB%82%98%EC%9D%B4%ED%94%84"
    option_txt = "실버다크브라운 포크1P&나이프1P"
    driver.get(url)

    # driver.execute_script('goCart2()')
    # alert = driver.switch_to.alert
    # alert.accept()
    # alert.dismiss()

    # driver.find_element('xpath', "//*select[@id='_goods_options']/option[text()={}]").format(option).click()
    option = Select(driver.find_element('xpath', "//*[@id='_goods_options']"))
    option.select_by_visible_text(option_txt)


def cart():

    # driver.get("https://www.daisomall.co.kr/member/login.php?url=")
    # driver.find_element('name', 'id').send_keys('sstlabs')
    # driver.find_element('name', 'pw').send_keys('Sstlabs1')
    # driver.find_element('name', 'pw').send_keys(Keys.ENTER)
    # sleep(5)


    for r in range(5, row_max):
        url = ws.cell(row=r, column=10).value
        # option = ws.cell(row=r, column=11).value
        num = ws.cell(row=r, column=5).value

        driver.get(url)

        try:

            # if option!=None:
            #     optionBox = Select(driver.find_element('xpath', "//*[@id='_goods_options']"))
            #     optionBox.select_by_visible_text(option)

            driver.execute_script('goCart2()')

            alert = driver.switch_to.alert
            alert.accept()
            alert.dismiss()
        except:
            error.append(num)
            print(num)
            continue

        

cart()
print(error)